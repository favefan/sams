from django.shortcuts import render, get_object_or_404, reverse
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse, FileResponse, StreamingHttpResponse
from django.db import models
from django.db.models import Q
from dashboard.models import User, Student, Activity, Entrylist
from django.contrib.auth.models import Group, Permission
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, permission_required
from django.utils import timezone
from django.core import serializers
import json
from openpyxl import *
import os,sys
from tempfile import NamedTemporaryFile
import codecs

login_page = 'dashboard/login.html'
account_manager_page = 'dashboard/account_manager.html'
index_page = 'dashboard/index.html'
name_or_pwd_error = '用户名或密码不正确'
user_already_exists = '用户名已存在'
warning_null_value = '用户名或密码不能为空'
sign_up_success = 'Sign up success! Please sign in.'

###Instrumental function
# 权限组的初始化
def get_or_create_group(group_name, codename_list):
    try:
        group = Group.objects.get(name=group_name)
    except Group.DoesNotExist:
        permissions = Permission.objects.filter(codename__in=codename_list)
        group = Group.objects.create(name=group_name)
        group.permissions.set(permissions)
        group.save()
    return group

# 网页显示message, 5s后跳转到red_url
def auto_red_html(message, red_url):
    return  message + \
            '<head>\
                <meta http-equiv="refresh" content="5;URL=' + red_url + '">\
            </head>\
            <body>\
                <p>5秒后自动跳转...<br>或点击\
                    <a href="' + red_url + '">此处</a>\
                直接跳转。</p>\
            </body>'

# 学生报名活动数与活动报名学生数计数更新
def entry_count_update(student, activity):
    act_count = Entrylist.objects.filter(student=student).count()
    stu_count = Entrylist.objects.filter(activity=activity).count()
    student.entry_count = act_count
    student.save()
    activity.entry_count = stu_count
    activity.save()
    return 

###
# 登录
def log_in(request):
    if request.method == 'POST':
        if request.POST:
            username = request.POST['uname']
            password = request.POST['pwd']
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                #return render(request, centre_page)
                return HttpResponseRedirect(reverse('dashboard:index'))
            else:
                return render(request, login_page, {'info': name_or_pwd_error})
        else:
            return render(request, login_page, {'info': warning_null_value})
    else:
        return render(request, login_page)

# 登出
def log_out(request):
    logout(request)
    return HttpResponseRedirect(reverse('dashboard:login'))

# 登录和未登录情况下不同的网站主页初始化
@login_required
def index(request):
    current_user = request.user
    if current_user.account is None:
        return HttpResponseRedirect(reverse('dashboard:open_acts'))
    else:
        scmark = 0
        thmark = 0
        student = current_user.account
        mine_entry_list = Entrylist.objects.filter(student=student)
        my_act_count = mine_entry_list.count()
        for one in mine_entry_list:
            if '第二课堂' in one.score_kind:
                scmark = scmark + one.score
            if '思想品德' in one.score_kind:
                thmark = thmark + one.score

        return render(request, 'dashboard/index.html', {'mine_entry_list': mine_entry_list, 'act_count': my_act_count, 'scmark': scmark, 'thmark': thmark })

# 学生管理页面
@login_required
@permission_required('dashboard.add_student', raise_exception=True)
def stu_manager(request):
    stus_list = Student.objects.all()
    return render(request, 'dashboard/stu_manager.html', { 'stus_list': stus_list })

# 账户管理页面
@login_required
@permission_required('dashboard.add_user', raise_exception=True)
def account_manager(request):
    accounts_list = User.objects.all()
    return render(request, 'dashboard/account_manager.html', { 'accounts_list': accounts_list })

# 活动管理页面
@login_required
@permission_required('dashboard.add_activity', raise_exception=True)
def act_manager(request):
    organizer = User.objects.get(id=request.session.get('_auth_user_id'))
    if organizer.is_superuser:
        mine_acts_list = Activity.objects.all()
    else:
        mine_acts_list = Activity.objects.filter(organizer=organizer)
    return render(request, 'dashboard/act_manager.html', { 'mine_acts_list': mine_acts_list })

# 信息修改页面
@login_required
def info_edit(request):
    return render(request, 'dashboard/info_edit.html')

# 获取公开活动
@login_required
def open_acts(request):
    open_acts_list = Activity.objects.all()
    return render(request, 'dashboard/open_acts.html', { 'open_acts_list': open_acts_list })

# 创建学生对象(自动创建对象对应账号)
@login_required
@permission_required('dashboard.add_student', raise_exception=True)
def create_student(request):
    codename_list = [ 
        'view_activity', 
        'add_entrylist', 
    ]
    normal_users_group = get_or_create_group('normal_users', codename_list)
    if request.POST:
        stu_id = request.POST['stu_id']
        stu_name = request.POST['stu_name']
        sex = request.POST['sex']
        depart = request.POST['depart']
        major = request.POST['major']
        en_year = request.POST['en_year']
        school_year = request.POST['school_year']
        username = stu_id
        password = None
        if len(stu_id) >= 6:
            password = stu_id[-6:]
        else:
            password = stu_id
        if Student.objects.filter(student_ID=stu_id).exists():
            return HttpResponse('输入的学号已存在，请检查.')
        else:
            student = Student(\
                student_ID=int(stu_id),\
                    name=stu_name,\
                        sex=int(sex),\
                            department=int(depart),\
                                major=major,\
                                    enroll_year=int(en_year),\
                                        schooling_year=int(school_year)\
                                            )
            if User.objects.filter(username=stu_id).exists():
                student.account_status = '账户冲突'
                student.save()
                return HttpResponse('学生信息创建成功, 但学生账户已经存在，无法关联账号，请检查。')
            else:
                student.save()
                new_user = User.objects.create_user(username=username, password=password, first_name=stu_name, account=student)
                new_user.save()
                normal_users_group.user_set.add(new_user)
                student.account_status = '已创建'
                student.save()
            return HttpResponse('学生信息创建成功，学生账户默认用户名为学号，默认密码为学号至少倒数六位。')
            #return render(request, centre_page, {'result_for_cs': 'Student and login account create success.username is: ' + username + ' , and password is: ' + password })
        #return render(request, centre_page, { 'result': [stu_id, stu_name, sex, depart, major, en_year, school_year] })

# 创建账号
@login_required
@permission_required('dashboard.add_user', raise_exception=True)
def create_account(request):
    # activity_admins_group = Group.objects.get_or_create(name='activity_admins', permissions=permissions)
    codename_list = [
        'add_activity', 
        'change_activity', 
        'view_activity', 
        'add_entrylist', 
        'change_entrylist', 
        'delete_entrylist', 
        'view_entrylist', 
        'view_student',
    ]
    activity_admins_group = get_or_create_group('activity_admins', codename_list)
    if request.method == 'POST':
        if request.POST:
            username = request.POST['uname']
            password = request.POST['pwd']
            if User.objects.filter(username=username).exists():
                return render(request, account_manager_page, {'info': user_already_exists})
            else:
                new_user = User.objects.create_user(username=username, password=password)
                new_user.save()
                activity_admins_group.user_set.add(new_user)
                #return render(request, sign_in_page, {'info': sign_up_success})
                return HttpResponse('账户创建成功.')
        else:
            return HttpResponse(warning_null_value)
    else:
        return render(request, account_manager_page)

# 创建活动
@login_required
@permission_required('dashboard.add_activity', raise_exception=True)
def create_activity(request):
    if request.POST:
        act_name = request.POST['act_name']
        print(act_name)
        intro = request.POST['intro']
        organizer = User.objects.get(id=request.session.get('_auth_user_id')) 
        create_date = timezone.now()
        capacity = request.POST['capacity']
        start_date = request.POST['start_date']
        end_date = request.POST['end_date']
        # if Activity.objects.filter(name=act_name).exists():
        activity = Activity(\
            name=act_name,\
                introduction=intro,\
                    organizer=organizer,\
                        create_date=create_date,\
                            capacity=int(capacity),\
                                start_date=start_date,\
                                    end_date=end_date\
                                        )
        activity.save()
        return HttpResponse('活动创建成功。')
    else:
        return HttpResponse('输入内容不能为空。')
    #return HttpResponseRedirect(reverse('dashboard:index'))
    #return render(request, centre_page, { 'result_for_ca': [act_name, intro, organizer, create_date, capacity, start_date, end_date]})

# 删除活动
@login_required
@permission_required('dashboard.delete_activity', raise_exception=True)
def delete_activity(request, activity_ID):
    act = Activity.objects.filter(activity_ID=activity_ID)
    if act.exists():
        act.delete()
    return HttpResponseRedirect(reverse('dashboard:index'))

# 删除学生对象
@login_required
@permission_required('dashboard.delete_student', raise_exception=True)
def delete_student(request, student_ID):
    stu = Student.objects.filter(student_ID=student_ID)
    stu_account = User.objects.filter(account=stu[0])
    if stu.exists():
        stu.delete()
        if stu_account.exists():
            stu_account.delete()
    return HttpResponseRedirect(reverse('dashboard:stu_manager'))

# 删除账号(关联了学生账号需要删除学生对象)
@login_required
@permission_required('dashboard.delete_user', raise_exception=True)
def delete_account(request, account_ID):
    account = User.objects.filter(id=account_ID)
    if account.exists():
        if account[0].account:
            return HttpResponse('此账号关联了学生对象，无法删除，若要删除请删除对应学生对象。')
        else:
            account.delete()
            return HttpResponseRedirect(reverse('dashboard:account_manager'))

# 账号修改
@login_required
def edit_account(request):
    account_id = request.POST['id']
    try:
        account = User.objects.get(id=account_id)
    except User.DoesNotExist:
        return HttpResponse('Ops!Fake id.')
    else:
        if 'first_name' in request.POST:
            first_name = request.POST['first_name']
            account.first_name = first_name
            account.save()
            return HttpResponse('姓名修改成功！')
        elif 'password' in request.POST:
            password = request.POST['password']
            account.set_password(password)
            account.save()
            return HttpResponse('密码修改成功！')
        else:
            return HttpResponse('Ops!Nothing be submitted.')

# 报名请求
@login_required
def enroll(request, activity_ID):
    current_user = request.user
    activity = Activity.objects.get(activity_ID=activity_ID)
    student = current_user.account
    entry_check = Entrylist.objects.filter(student=student, activity=activity)
    if entry_check.exists():
        return HttpResponse(auto_red_html('你已经报名了，请不要重复报名', '/dashboard/open_acts/'))
    else:
        enroll_activity = Entrylist(student=student, activity=activity, entry_date=timezone.now())
        enroll_activity.save()
        entry_count_update(student, activity)
        return HttpResponseRedirect(reverse('dashboard:open_acts'))

# 活动信息
@login_required
@permission_required('dashboard.change_activity', raise_exception=True)
def act_info(request, activity_ID):
    activity = Activity.objects.get(activity_ID=activity_ID)
    entry_list = Entrylist.objects.filter(activity=activity)
    return render(request, 'dashboard/act_info.html', { 'act': activity, 'entry_list': entry_list, 'count': entry_list.count() })

# 学生信息
@login_required
@permission_required('dashboard.change_student', raise_exception=True)
def stu_info(request, student_ID):
    scmark = 0
    thmark = 0
    student = Student.objects.get(student_ID=student_ID)
    mine_entry_list = Entrylist.objects.filter(student=student)
    my_act_count = mine_entry_list.count()
    for one in mine_entry_list:
        if '第二课堂' in one.score_kind:
            scmark = scmark + one.score
        if '思想品德' in one.score_kind:
            thmark = thmark + one.score
    return render(request, 'dashboard/stu_info.html', { 'student':student, 'mine_entry_list': mine_entry_list, 'act_count': my_act_count, 'scmark': scmark, 'thmark': thmark })

# 活动编辑
@login_required
@permission_required('dashboard.change_activity', raise_exception=True)
def edit_act(request, activity_ID):
    activity = Activity.objects.get(activity_ID=activity_ID)
    name = request.POST['act_name']
    intro = request.POST['intro']
    capacity = request.POST['capacity']
    start_date = request.POST['start_date']
    end_date = request.POST['end_date']
    if name != activity.name:
        activity.name = name
    if intro != activity.introduction:
        activity.introduction = intro
    if capacity != activity.capacity:
        activity.capacity = capacity
    if start_date != activity.start_date:
        activity.start_date = start_date
    if end_date != activity.end_date:
        activity.end_date = end_date
    activity.save()
    return HttpResponse('修改成功!')    

# 活动搜索功能    
@login_required
def search(request):
    search_content = request.POST['search_content']
    result = Student.objects.filter(
        Q(student_ID__icontains=search_content)|
        Q(name__icontains=search_content))
    if result.exists():
        data = serializers.serialize('json', result)
        return HttpResponse(data, content_type="application/json;charset=utf-8")
    else:
        return HttpResponse('0')

# 手动添加报名信息
@login_required
def manual_enroll(request, activity_ID, student_ID):
    activity = Activity.objects.get(activity_ID=activity_ID)
    student = Student.objects.get(student_ID=student_ID)
    entry_check = Entrylist.objects.filter(student=student, activity=activity)
    if entry_check.exists():
        return HttpResponse('002')
    else:
        enroll_activity = Entrylist(student=student, activity=activity, entry_date=timezone.now())
        enroll_activity.save()
        entry_count_update(student, activity)
        return HttpResponse('200')

# 奖项/学分设置
@login_required
def award_give(request):
    act_id = request.POST['act_id']
    stu_id = request.POST['stu_id']
    award = request.POST['award']
    score_kind = request.POST['score_kind']
    score = request.POST['score']
    activity = Activity.objects.get(activity_ID=act_id)
    student = Student.objects.get(student_ID=stu_id)
    entry = Entrylist.objects.get(student=student, activity=activity)
    if entry:
        entry.awards = award
        entry.score_kind = score_kind
        entry.score = score
        entry.save()
        return HttpResponse('奖项设置成功！')
    else:
        return HttpResponse('奖项设置失败，请联系管理员！')

# 活动分享
def shared(request, activity_ID):
    activity = Activity.objects.get(activity_ID=activity_ID)
    count = Entrylist.objects.filter(activity=activity).count()
    return render(request, 'dashboard/shared.html', { 'act': activity, 'count': count })

# #
def any_login(request):
    if request.method == 'POST':
        if request.POST:
            username = request.POST['uname']
            password = request.POST['pwd']
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return HttpResponse('登陆成功！')
            else:
                return HttpResponse(name_or_pwd_error)
        else:
            return HttpResponse(warning_null_value)
    else:
        return HttpResponse('Ops!')

# 导出活动报名名单
def get_report(request, activity_ID):
    tno = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    title = ['序号', '姓名', '学号', '班级', '获得奖项', '加分类型', '分数']
    activity = Activity.objects.get(activity_ID=activity_ID)
    entry_list = Entrylist.objects.filter(activity=activity)
    wb = Workbook()
    ws = wb.active
    i = 0
    for t in tno:
        ws[t + '1'] = title[i]
        i = i + 1
    i = 2
    for one in entry_list:
        ws['A' + str(i)] = i - 1
        ws['B' + str(i)] = one.student.name
        ws['C' + str(i)] = str(one.student.student_ID)
        ws['D' + str(i)] = str(one.student.enroll_year) + '级' + one.student.major + str(one.student.department) + '班'
        ws['E' + str(i)] = one.awards
        ws['F' + str(i)] = one.score_kind
        ws['G' + str(i)] = one.score
        i = i + 1
    file_name = activity.name
    wb.save(os.path.abspath('./dashboard/static/dashboard/upload/' + str(activity_ID) + '.xlsx'))
    # file = codecs.open(str(activity_ID) + '.xlsx', 'rb', encoding= u'utf-8', errors='ignore')
    # response = StreamingHttpResponse(file)
    # response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    # response['Content-Disposition'] = 'attachment;filename=活动' + file_name + '的报名详情.xlsx'

    return HttpResponse('/static/dashboard/upload/' + str(activity_ID) + '.xlsx')

# 数据可视化
def graph_data(request):
    activity_name_list = []
    activity_entry_count_list = []
    student_name_list = []
    student_entry_count_list = []
    activitys = Activity.objects.order_by('-entry_count')[:3]
    students = Student.objects.order_by('-entry_count')[:3]
    for act in activitys:
        activity_name_list.append(act.name)
        activity_entry_count_list.append(act.entry_count)
    for stu in students:
        student_name_list.append(stu.name)
        student_entry_count_list.append(stu.entry_count)
    return render(request, 'dashboard/graph_data.html', { 'activity_name_list': activity_name_list, 'activity_entry_count_list': activity_entry_count_list, 'student_name_list':student_name_list, 'student_entry_count_list':student_entry_count_list })